from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_BG   = "1F4E79"
CAT_COLORS  = {
    "Launch Cost":                    "D6E4F0",
    "Starlink Mass & Cost":           "D9EAD3",
    "Reflect Orbital – Satellite Specs": "FCE5CD",
    "Reflect Orbital – Mission":      "EAD1DC",
    "Data Quality Notes":             "F3F3F3",
}
LINK_FONT   = Font(color="1155CC", underline="single", name="Arial", size=9)
PLAIN_FONT  = Font(name="Arial", size=9)
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)

data = [
    ("Launch Cost", "F9 dedicated list price: $74M, 22,800 kg → $3,246/kg",
     "Upper bound. What Reflect actually pays is the rideshare rate.",
     [("Wikipedia – Falcon 9", "https://en.wikipedia.org/wiki/Falcon_9")]),

    ("Launch Cost", "F9 rideshare (Transporter SSO): $6,000/kg",
     "What Reflect pays today. Nowack quoted ~$6,500/kg in Monocle interview.",
     [("Wikipedia – Falcon 9", "https://en.wikipedia.org/wiki/Falcon_9"),
      ("Monocle – Nowack interview", "https://monocle.com/business/aviation/reflect-orbital-aerospace-startup/")]),

    ("Launch Cost", "F9 internal marginal cost: ~$15M/launch → ~$857/kg",
     "SpaceX's actual cost. Floor on how cheap F9 could ever get for Reflect.",
     [("Wikipedia – Falcon 9", "https://en.wikipedia.org/wiki/Falcon_9")]),

    ("Launch Cost", "Starship aspirational: $10M/launch, 150,000 kg → $67/kg",
     "SpaceX stated target at high cadence. Not achieved. The number that makes Reflect's economics work.",
     [("SpaceNexus – launch cost economics", "https://spacenexus.us/blog/economics-satellite-launch-cost-per-kilogram")]),

    ("Launch Cost", "Starship conservative near-term: ~$30M/launch → ~$200/kg",
     "More realistic early operational estimate from analysts.",
     [("SpaceNexus – launch cost economics", "https://spacenexus.us/blog/economics-satellite-launch-cost-per-kilogram")]),

    ("Starlink Mass & Cost", "v1.0: 260 kg, $200k/unit → $769/kg",
     "First operational generation. Cost from Quilty Space analyst estimate via SpaceNews.",
     [("Mass – Skyrocket", "https://space.skyrocket.de/doc_sdat/starlink-v1-0.htm"),
      ("Cost – SpaceNews", "https://spacenews.com/starlink-soars-spacexs-satellite-internet-surprises-analysts-with-6-6-billion-revenue-projection/")]),

    ("Starlink Mass & Cost", "v1.5: ~305 kg, ~$250k/unit (estimated) → ~$820/kg",
     "Added laser ISLs. Mass from Teslarati/FCC filing. Cost interpolated — no primary source.",
     [("Mass – Teslarati", "https://www.teslarati.com/spacex-unveils-next-gen-starlink-v2-mini-satellites-ahead-of-monday-launch/"),
      ("Mass – Wikipedia Starlink", "https://en.wikipedia.org/wiki/Starlink"),
      ("Cost – interpolated (no source)", None)]),

    ("Starlink Mass & Cost", "v2 Mini: 800 kg, $800k/unit → $1,000/kg",
     "Current generation, launched on F9. Cost from Quilty Space via SpaceNews.",
     [("Mass – Wikipedia Starlink", "https://en.wikipedia.org/wiki/Starlink"),
      ("Cost – SpaceNews", "https://spacenews.com/starlink-soars-spacexs-satellite-internet-surprises-analysts-with-6-6-billion-revenue-projection/")]),

    ("Starlink Mass & Cost", "v2 full (Starship-only): ~1,250 kg, ~$1M/unit (estimated) → ~$800/kg",
     "Too large for F9, not yet launched. Mass from Musk statement. Cost estimated — no primary source.",
     [("Mass – Teslarati", "https://www.teslarati.com/spacex-elon-musk-next-gen-starlink-satellite-details/"),
      ("Cost – estimate only", None)]),

    ("Starlink Mass & Cost", "v3 (projected): 1,500 kg, $1.2M/unit → $800/kg",
     "Future generation. Both figures from Quilty Space via SpaceNews.",
     [("SpaceNews – Quilty Space", "https://spacenews.com/starlink-soars-spacexs-satellite-internet-surprises-analysts-with-6-6-billion-revenue-projection/")]),

    ("Reflect Orbital – Satellite Specs", "Eärendil-1 mirror mass: 16 kg",
     "Mirror only (Mylar film on deployable booms), not total satellite.",
     [("Wikipedia – Reflect Orbital", "https://en.wikipedia.org/wiki/Reflect_Orbital"),
      ("Space.com", "https://www.space.com/orbiting-mirror-boost-solar-power-production")]),

    ("Reflect Orbital – Satellite Specs", "Eärendil-1 total satellite mass: 100 kg",
     'Full sat incl. bus, propulsion, avionics. From Monocle interview: "100 kg packed into a box the size of a microwave".',
     [("Monocle – Nowack interview", "https://monocle.com/business/aviation/reflect-orbital-aerospace-startup/"),
      ("Orbital Today", "https://orbitaltoday.com/2025/07/31/startup-plans-to-beam-sunlight-to-earth-using-space-mirrors/")]),

    ("Reflect Orbital – Satellite Specs", "Eärendil-1 mirror area: 324 m² (18m × 18m)",
     "Derived. m²/kg = 20.3 (mirror only) / 3.24 (whole satellite).",
     [("Wikipedia – Reflect Orbital", "https://en.wikipedia.org/wiki/Reflect_Orbital")]),

    ("Reflect Orbital – Satellite Specs", "Future constellation mirror: 54m × 55m (~2,916–3,025 m²)",
     "Stated goal for production satellites. ~9× more area than Eärendil.",
     [("Payload Space – Series A", "https://payloadspace.com/reflect-orbital-raises-20m-series-a/"),
      ("LiveScience", "https://www.livescience.com/space/space-exploration/controversial-startups-plan-to-sell-sunlight-using-giant-mirrors-in-space-would-be-catastrophic-and-horrifying-astronomers-warn")]),

    ("Reflect Orbital – Satellite Specs", "Future constellation satellite total mass: NO PUBLIC DATA",
     "Estimate: mirror film ~144 kg (9× Eärendil), bus ~200–400 kg → total ~300–500 kg. Treat as unknown.",
     []),

    ("Reflect Orbital – Mission", "Orbit: SSO, 600–650 km altitude",
     "Sets ground track geometry and pass duration.",
     [("Wikipedia – Reflect Orbital", "https://en.wikipedia.org/wiki/Reflect_Orbital")]),

    ("Reflect Orbital – Mission", "Beam diameter on ground: 5 km",
     "Set by sun's angular diameter (0.53°) projected from altitude. Physics constraint, not a design choice.",
     [("Wikipedia – Reflect Orbital", "https://en.wikipedia.org/wiki/Reflect_Orbital")]),

    ("Reflect Orbital – Mission", "Target irradiance: 200 W/m²",
     "Reflect's stated goal for the full constellation. ~20% of midday sun.",
     [("Monocle – Nowack interview", "https://monocle.com/business/aviation/reflect-orbital-aerospace-startup/")]),

    ("Reflect Orbital – Mission", "Pricing: ~$5,000/hour per satellite",
     "Nowack's stated figure for lighting customers. Solar farm revenue-share also discussed, no public rate.",
     [("Futurism", "https://futurism.com/space/fcc-huge-mirror-satellite")]),
]

quality_notes = [
    "All Starlink cost figures are analyst estimates (Quilty Space), not SpaceX disclosures. Treat as ±30%.",
    "Starlink mass figures are well-sourced (FCC filings, NSF launch reports, Wikipedia citing primary sources).",
    "Reflect total satellite mass (100 kg) comes from one interview — treat as approximate.",
    "Future constellation satellite mass has no source and must be modeled or assumed.",
    "F9 rideshare pricing is publicly listed and reliable.",
    "Starship cost figures are targets, not achieved prices.",
]

wb = Workbook()
ws = wb.active
ws.title = "Sources"

thin = Side(style="thin", color="CCCCCC")
border = Border(bottom=Side(style="thin", color="CCCCCC"))

# Header row
headers = ["Category", "Item / Fact", "Notes", "Source 1", "Source 2", "Source 3"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = PatternFill("solid", start_color=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.row_dimensions[1].height = 20

def fill_for(cat):
    color = CAT_COLORS.get(cat, "FFFFFF")
    return PatternFill("solid", start_color=color)

def write_link(cell, label, url):
    cell.value = label
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT
    else:
        cell.font = Font(color="888888", name="Arial", size=9, italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

for row_idx, (cat, item, notes, sources) in enumerate(data, start=2):
    fill = fill_for(cat)
    vals = [cat, item, notes]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row_idx, column=col, value=v)
        c.font = PLAIN_FONT
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top")

    for src_col, (label, url) in enumerate(sources[:3], start=4):
        c = ws.cell(row=row_idx, column=src_col)
        c.fill = fill
        write_link(c, label, url)

    for src_col in range(len(sources) + 4, 7):
        c = ws.cell(row=row_idx, column=src_col)
        c.fill = fill

    ws.row_dimensions[row_idx].height = 40

# Data quality notes section
note_start = len(data) + 3
c = ws.cell(row=note_start - 1, column=1, value="Data Quality Notes")
c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
c.fill = PatternFill("solid", start_color="555555")
ws.merge_cells(start_row=note_start - 1, start_column=1, end_row=note_start - 1, end_column=6)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[note_start - 1].height = 18

fill_note = fill_for("Data Quality Notes")
for i, note in enumerate(quality_notes):
    r = note_start + i
    c = ws.cell(row=r, column=1, value=note)
    c.font = PLAIN_FONT
    c.fill = fill_note
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28

# Column widths
col_widths = [28, 42, 48, 32, 32, 32]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Freeze header
ws.freeze_panes = "A2"

out = "C:/_Dev/charts-and-data-analysis/Reflect-Orbital/sso-land-proximity/data/reflect_orbital_sources.xlsx"
wb.save(out)
print(f"Saved: {out}")
