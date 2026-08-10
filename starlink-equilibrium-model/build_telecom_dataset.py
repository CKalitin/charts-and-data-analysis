"""Build data/telecom_market_by_country.csv from bulk public sources.

Phase 1 of the Starlink equilibrium model (see CLAUDE.md). Merges:
  - World Bank API: population, internet-user %, fixed-broadband/100, mobile/100,
    GNI per capita PPP, GDP per capita PPP (all "most recent non-empty value" per
    country, mrnev=1). GDP per capita added 2026-08-09 alongside the existing GNI
    figure specifically so affordability analysis can compare the two -- they
    diverge materially for countries with large multinational profit-shifting (GDP
    >> GNI, e.g. Ireland, Luxembourg) or large inbound remittances/outbound resource
    income (GNI can exceed or trail GDP the other way). The international
    affordability-target literature (UN Broadband Commission, A4AI "1 for 2" / "Journey
    from 1 to 5") standardizes on GNI per capita specifically to avoid this distortion
    -- see charts/affordability.py for the applied comparison.
  - bestbroadbanddeals.co.uk worldwide mobile data pricing survey (2023 vintage,
    237 countries) -> average $/GB for 1GB of mobile data
  - broadband.co.uk global broadband price league (Feb 2026, 214 countries) ->
    average fixed-broadband $/month

Two DERIVED fields use an explicit, documented usage-volume assumption to convert
between $/month and $/GB (no per-country usage-volume dataset exists at this
granularity) — see the "illustrative assumption" columns and telecom_market_by_country.md.

Legacy satellite ISP presence is a hand-coded categorical flag, not a sourced number
(no bulk country-level source exists for this — see research notes in CLAUDE.md).

Run: python build_telecom_dataset.py
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl

DATA = Path(__file__).resolve().parent / "data"
RAW = DATA / "raw"

# Illustrative, DOCUMENTED usage-volume assumptions used only to cross-derive
# $/GB <-> $/month where the bulk source only gave us one side. Not per-country data.
ASSUMED_FIXED_BROADBAND_GB_PER_MONTH = 300.0   # global illustrative household usage
ASSUMED_MOBILE_GB_PER_MONTH = 10.0             # global illustrative mobile usage

# Coverage-gap correction assumption (2026-08-09, user-requested) -- see the inline
# comment at its point of use for why this exists and its documented limitations.
HIGH_INCOME_COVERAGE_FLOOR_PCT = 97.0

# Country-name aliasing: World Bank name -> alternate spellings seen in the other
# two sources. Built by inspecting unmatched names after a first merge pass.
NAME_ALIASES = {
    "united states": "united states",
    "russian federation": "russia",
    "korea, rep.": "south korea",
    "korea, dem. people's rep.": "north korea",
    "iran, islamic rep.": "iran",
    "egypt, arab rep.": "egypt",
    "venezuela, rb": "venezuela",
    "vietnam": "viet nam",
    "lao pdr": "laos",
    "congo, dem. rep.": "dr congo",
    "congo, rep.": "congo",
    "cote d'ivoire": "ivory coast",
    "czechia": "czech republic",
    "slovak republic": "slovakia",
    "kyrgyz republic": "kyrgyzstan",
    "brunei darussalam": "brunei",
    "syrian arab republic": "syria",
    "yemen, rep.": "yemen",
    "turkiye": "turkey",
    "gambia, the": "gambia",
    "bahamas, the": "bahamas",
    "st. lucia": "saint lucia",
    "st. vincent and the grenadines": "saint vincent and the grenadines",
    "st. kitts and nevis": "saint kitts and nevis",
    "cabo verde": "cape verde",
    "eswatini": "eswatini",
    "north macedonia": "north macedonia",
    "myanmar": "myanmar",
    "lao pdr": "laos",
    "hong kong sar, china": "hong kong",
    "macao sar, china": "macau",
    "united arab emirates": "united arab emirates",
    "united kingdom": "united kingdom",
    "somalia, fed. rep.": "somalia",
    "congo, rep.": "republic of the congo",
    "micronesia, fed. sts.": "micronesia",
    "puerto rico (us)": "puerto rico",
    "virgin islands (u.s.)": "united states virgin islands",
    "sint maarten (dutch part)": "sint maarten",
}


def norm(name: str) -> str:
    n = name.strip().lower().replace("‑", "-")  # non-breaking hyphen -> plain hyphen
    n = n.replace("â€‘", "-")     # mojibake hyphen seen in broadband.co.uk sheet
    n = n.replace(".", "").replace(",", "")
    return NAME_ALIASES.get(name.strip().lower(), n)


def load_wb_countries():
    d = json.load(open(RAW / "wb_countries.json", encoding="utf-8"))[1]
    return [c for c in d if c["region"]["id"] != "NA"]


def load_wb_indicator(code: str):
    d = json.load(open(RAW / f"wb_{code}.json", encoding="utf-8"))[1]
    out = {}
    for r in d:
        if r["value"] is not None and r["countryiso3code"]:
            out[r["countryiso3code"]] = r["value"]
    return out


def load_mobile_gb_pricing():
    """ISO2 -> avg USD per 1GB mobile data (2023 survey, bestbroadbanddeals.co.uk)."""
    wb = openpyxl.load_workbook(RAW / "mobile_data_pricing_2023.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    out = {}
    for r in rows:
        iso2, usd_per_gb = r[1], r[8]
        if iso2 and usd_per_gb is not None:
            out[iso2] = float(usd_per_gb)
    return out


def load_broadband_month_pricing():
    """normalized country name -> avg USD/month fixed broadband (Feb 2026, broadband.co.uk)."""
    wb = openpyxl.load_workbook(RAW / "broadband_pricing_2026.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    out = {}
    for r in rows:
        name, usd = r[1], r[7]
        if name and usd is not None:
            out[norm(name)] = float(usd)
    return out


# Hand-coded, NOT bulk-sourced: legacy GEO satellite broadband ISP presence.
# Coarse categorical judgment from public knowledge (Viasat/HughesNet/Eutelsat
# Konnect/NBN Sky Muster coverage footprints) -- treat as approximate, see MD.
SATELLITE_ISP_PRESENT_ISO3 = {
    # North America -- Viasat/HughesNet both operate at meaningful rural scale
    "USA", "CAN", "MEX",
    # Europe -- Eutelsat Konnect / Viasat regional footprint
    "GBR", "FRA", "DEU", "ITA", "ESP", "POL", "ROU", "GRC", "PRT", "IRL",
    # Australia -- NBN Sky Muster is a national fixed-satellite program
    "AUS", "NZL",
    # Brazil -- Viasat/Hughesnet Brazil operate commercially
    "BRA",
    # India -- Hughes Communications India JV
    "IND",
}


def main():
    countries = load_wb_countries()
    pop = load_wb_indicator("SP.POP.TOTL")
    netuser = load_wb_indicator("IT.NET.USER.ZS")
    fbb = load_wb_indicator("IT.NET.BBND.P2")
    mobile_pen = load_wb_indicator("IT.CEL.SETS.P2")
    gni = load_wb_indicator("NY.GNP.PCAP.PP.CD")
    gdp = load_wb_indicator("NY.GDP.PCAP.PP.CD")
    mobile_gb = load_mobile_gb_pricing()
    bb_month = load_broadband_month_pricing()

    unmatched_mobile_gb = []
    unmatched_bb_month = []
    rows = []
    for c in countries:
        iso3, iso2, name = c["id"], c["iso2Code"], c["name"]
        region = c["region"]["value"].strip()
        income = c["incomeLevel"]["value"]

        p = pop.get(iso3)
        net_pct = netuser.get(iso3)
        fbb_p100 = fbb.get(iso3)
        mob_p100 = mobile_pen.get(iso3)
        gni_pc = gni.get(iso3)
        gdp_pc = gdp.get(iso3)

        mgb = mobile_gb.get(iso2)
        if mgb is None:
            unmatched_mobile_gb.append(name)

        bbm = bb_month.get(norm(name))
        if bbm is None:
            unmatched_bb_month.append(name)

        unconnected = None
        if p is not None and net_pct is not None:
            unconnected = round(p * (1 - net_pct / 100.0))

        # Coverage-gap correction (added 2026-08-09, user-requested): World Bank's
        # internet_user_pct is a USAGE-gap metric (counts non-use for ANY reason --
        # age, digital literacy, personal choice -- not just lack of infrastructure).
        # High-income countries have near-universal terrestrial coverage almost by
        # definition of that income classification, so treating their full usage
        # gap as Starlink-addressable "unconnected" population overstates the
        # coverage gap (the specific finding flagged in Phase 1 -- France ~11%,
        # Italy ~11% "unconnected" despite full coverage). No country-level bulk
        # coverage-gap dataset was found to be accessible (ITU DataHub blocks
        # automated fetch, GSMA's export links are dead -- see telecom_market_by_country.md)
        # so this is a documented, targeted assumption, not measured data: for
        # High income countries only, floor the effective connected % at
        # HIGH_INCOME_COVERAGE_FLOOR_PCT (assumed near-universal infrastructure
        # coverage), leaving lower income tiers unchanged since usage and coverage
        # gaps are more plausibly correlated there.
        unconnected_corrected = None
        if p is not None and net_pct is not None:
            floor = HIGH_INCOME_COVERAGE_FLOOR_PCT if income == "High income" else 0.0
            effective_connected_pct = max(net_pct, floor)
            unconnected_corrected = round(p * (1 - effective_connected_pct / 100.0))

        cellular_dominant = None
        if fbb_p100 is not None and mob_p100 is not None:
            # Fixed broadband penetration far below mobile penetration -> mobile-first market.
            cellular_dominant = (fbb_p100 < 5.0) or (mob_p100 > 0 and fbb_p100 / mob_p100 < 0.08)

        bb_gb_price = (bbm / ASSUMED_FIXED_BROADBAND_GB_PER_MONTH) if bbm is not None else None
        mobile_month_price = (mgb * ASSUMED_MOBILE_GB_PER_MONTH) if mgb is not None else None

        connectivity_cost_pct_gni = None
        if bbm is not None and gni_pc is not None and gni_pc > 0:
            connectivity_cost_pct_gni = round(100 * (bbm * 12) / gni_pc, 2)

        rows.append({
            "iso3": iso3,
            "country": name,
            "region": region,
            "income_level": income,
            "population": int(p) if p is not None else "",
            "internet_user_pct": round(net_pct, 1) if net_pct is not None else "",
            "unconnected_population_est": unconnected if unconnected is not None else "",
            "unconnected_population_est_coverage_corrected": unconnected_corrected if unconnected_corrected is not None else "",
            "fixed_broadband_subs_per_100": round(fbb_p100, 2) if fbb_p100 is not None else "",
            "mobile_subs_per_100": round(mob_p100, 1) if mob_p100 is not None else "",
            "cellular_dominant_market": cellular_dominant if cellular_dominant is not None else "",
            "fixed_broadband_usd_per_month": round(bbm, 2) if bbm is not None else "",
            "fixed_broadband_usd_per_gb_illustrative": round(bb_gb_price, 4) if bb_gb_price is not None else "",
            "mobile_usd_per_gb": round(mgb, 4) if mgb is not None else "",
            "mobile_usd_per_month_illustrative": round(mobile_month_price, 2) if mobile_month_price is not None else "",
            "gni_per_capita_ppp_usd": round(gni_pc) if gni_pc is not None else "",
            "gdp_per_capita_ppp_usd": round(gdp_pc) if gdp_pc is not None else "",
            "connectivity_cost_pct_of_gni": connectivity_cost_pct_gni if connectivity_cost_pct_gni is not None else "",
            "legacy_satellite_isp_present": iso3 in SATELLITE_ISP_PRESENT_ISO3,
        })

    fieldnames = list(rows[0].keys())
    out_path = DATA / "telecom_market_by_country.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    print(f"Wrote {len(rows)} rows to {out_path}")
    print(f"Unmatched mobile $/GB (no bestbroadbanddeals row): {len(unmatched_mobile_gb)}")
    print(f"  {unmatched_mobile_gb}")
    print(f"Unmatched broadband $/month (no broadband.co.uk row): {len(unmatched_bb_month)}")
    print(f"  {unmatched_bb_month}")


if __name__ == "__main__":
    main()
